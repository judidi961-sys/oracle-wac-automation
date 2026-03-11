"""
create_example_excel.py - Genera un archivo Excel de ejemplo para el bot WAC

Crea el archivo data/wac_update.xlsx con columnas correctas y datos de prueba.

Uso:
    python create_example_excel.py
    python create_example_excel.py --output data/mi_archivo.xlsx
"""

import argparse
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config


def create_example_excel(output_path: str = config.EXCEL_INPUT_FILE) -> None:
    """
    Crea un archivo Excel de ejemplo con columnas y datos de prueba.

    Args:
        output_path: Ruta donde se guardará el archivo.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WAC Update"

    # ── Header row ─────────────────────────────────────────────────
    headers = [
        config.COL_SKU,
        config.COL_LOCATION,
        config.COL_NEW_WAC,
        config.COL_UNITS,
        config.COL_CURRENT_WAC,
        config.COL_STATUS,
        config.COL_NOTES,
        config.COL_TIMESTAMP,
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # ── Sample data rows ────────────────────────────────────────────
    sample_rows = [
        ("SKU-001001", "1001", 125.50, 100),
        ("SKU-001002", "1002", 89.99,  250),
        ("SKU-002001", "2001", 45.75,  50),
        ("SKU-002005", "2005", 310.00, 10),
        ("SKU-003001", "3001", 15.30,  500),
    ]

    pending_fill = PatternFill("solid", fgColor="FFEB9C")
    for row_idx, (sku, location, new_wac, units) in enumerate(sample_rows, start=2):
        ws.cell(row=row_idx, column=1, value=sku)
        ws.cell(row=row_idx, column=2, value=location)
        ws.cell(row=row_idx, column=3, value=new_wac)
        ws.cell(row=row_idx, column=4, value=units)
        status_cell = ws.cell(row=row_idx, column=6, value=config.STATUS_PENDING)
        status_cell.fill = pending_fill

    # ── Column widths ───────────────────────────────────────────────
    column_widths = [15, 12, 14, 10, 14, 14, 40, 22]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(str(path))
    print(f"✅ Archivo Excel de ejemplo creado en: {path}")
    print(f"   Filas de datos: {len(sample_rows)}")
    print(f"   Columnas: {', '.join(headers)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crea un archivo Excel de ejemplo para el bot WAC."
    )
    parser.add_argument(
        "--output",
        default=config.EXCEL_INPUT_FILE,
        help=f"Ruta de salida (default: {config.EXCEL_INPUT_FILE})",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    create_example_excel(args.output)
