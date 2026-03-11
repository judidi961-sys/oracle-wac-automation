"""
src/excel_handler.py - Manejo de archivos Excel con openpyxl

Lee los registros SKU desde el archivo de entrada y escribe de vuelta
el estado y notas de cada fila en tiempo real.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Generator

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger("wac_bot")

# Cell fill colours for status feedback
_FILL_SUCCESS = PatternFill("solid", fgColor="C6EFCE")  # green
_FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")    # red
_FILL_PENDING = PatternFill("solid", fgColor="FFEB9C")  # yellow


class ExcelHandler:
    """
    Gestiona la lectura y escritura del archivo Excel de trabajo.
    """

    def __init__(self, file_path: str | None = None) -> None:
        self._path = Path(file_path or config.EXCEL_INPUT_FILE)
        self._wb: Workbook | None = None
        self._ws: Worksheet | None = None
        self._header_map: dict[str, int] = {}

    # ──────────────────────────────────────────
    # Public API
    # ──────────────────────────────────────────

    def open(self) -> None:
        """Abre el archivo Excel y valida las columnas requeridas."""
        if not self._path.exists():
            raise FileNotFoundError(
                f"Archivo Excel no encontrado: {self._path}\n"
                "Ejecuta 'python create_example_excel.py' para crear uno de ejemplo."
            )

        self._wb = openpyxl.load_workbook(str(self._path))
        self._ws = self._wb.active
        self._build_header_map()
        self._validate_required_columns()
        logger.info("Archivo Excel abierto: %s (%d filas de datos)", self._path, self.data_row_count)

    def close(self) -> None:
        """Cierra el libro de trabajo sin guardar (usa save() para persistir)."""
        if self._wb:
            self._wb.close()
            self._wb = None
            self._ws = None
            logger.debug("Archivo Excel cerrado.")

    def save(self) -> None:
        """Guarda los cambios en el archivo Excel."""
        if self._wb:
            self._wb.save(str(self._path))
            logger.debug("Archivo Excel guardado: %s", self._path)

    def iter_rows(self) -> Generator[dict, None, None]:
        """
        Itera sobre las filas de datos (omite la cabecera).

        Yields:
            Diccionario con los datos de la fila más la clave '_row_number'.
        """
        if self._ws is None:
            raise RuntimeError("El archivo Excel no está abierto. Llama a open() primero.")

        for row_idx in range(2, self._ws.max_row + 1):
            row_data = self._read_row(row_idx)
            sku = row_data.get(config.COL_SKU, "")
            if not sku:
                continue  # Skip empty rows
            row_data["_row_number"] = row_idx
            yield row_data

    def update_row_status(
        self,
        row_number: int,
        status: str,
        notes: str = "",
        current_wac: float | None = None,
    ) -> None:
        """
        Escribe el resultado de procesamiento en la fila correspondiente.

        Args:
            row_number: Número de fila en la hoja (1-indexed, primera fila de datos = 2).
            status: Valor de estado (usar constantes STATUS_* de config).
            notes: Notas adicionales o mensaje de error.
            current_wac: Valor WAC actual leído de Oracle (opcional).
        """
        if self._ws is None:
            return

        fill = {
            config.STATUS_SUCCESS: _FILL_SUCCESS,
            config.STATUS_ERROR: _FILL_ERROR,
            config.STATUS_SKIPPED: _FILL_PENDING,
        }.get(status, _FILL_PENDING)

        self._write_cell(row_number, config.COL_STATUS, status, fill=fill)
        self._write_cell(row_number, config.COL_NOTES, notes)
        self._write_cell(row_number, config.COL_TIMESTAMP, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if current_wac is not None:
            self._write_cell(row_number, config.COL_CURRENT_WAC, current_wac)

        self.save()

    @property
    def data_row_count(self) -> int:
        """Número de filas de datos (excluye la cabecera)."""
        if self._ws is None:
            return 0
        return max(0, self._ws.max_row - 1)

    # ──────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────

    def _build_header_map(self) -> None:
        """Construye un mapa de nombre de columna → índice de columna (1-indexed)."""
        if self._ws is None:
            return
        self._header_map = {}
        for cell in self._ws[1]:
            if cell.value:
                self._header_map[str(cell.value).strip()] = cell.column

    def _validate_required_columns(self) -> None:
        """Verifica que las columnas requeridas existan en el encabezado."""
        missing = [col for col in config.REQUIRED_COLUMNS if col not in self._header_map]
        if missing:
            raise ValueError(
                f"Columnas requeridas no encontradas en el Excel: {missing}\n"
                f"Columnas encontradas: {list(self._header_map.keys())}"
            )

    def _read_row(self, row_idx: int) -> dict:
        """Lee una fila completa y devuelve un dict con los valores."""
        row: dict = {}
        for col_name, col_idx in self._header_map.items():
            cell = self._ws.cell(row=row_idx, column=col_idx)
            row[col_name] = cell.value
        return row

    def _write_cell(
        self,
        row_number: int,
        col_name: str,
        value,
        fill: PatternFill | None = None,
    ) -> None:
        """Escribe un valor en la celda indicada, añadiendo la columna si no existe."""
        if col_name not in self._header_map:
            # Add the column at the end
            next_col = (max(self._header_map.values()) + 1) if self._header_map else 1
            self._header_map[col_name] = next_col
            header_cell = self._ws.cell(row=1, column=next_col, value=col_name)
            header_cell.font = Font(bold=True)

        col_idx = self._header_map[col_name]
        cell = self._ws.cell(row=row_number, column=col_idx, value=value)
        if fill:
            cell.fill = fill
